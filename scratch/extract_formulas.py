import openpyxl

file_path = 'VAT N INVENTORY(SSA).xlsm'

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['Costing_Budget']
    
    print("--- Costing_Budget Formulas ---")
    # Print first few rows of formulas
    for row in sheet.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                print(f"Cell {cell.coordinate}: {cell.value}")
                
    sheet_p = wb['Purchase_Book']
    print("\n--- Purchase_Book Formulas ---")
    for row in sheet_p.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                print(f"Cell {cell.coordinate}: {cell.value}")

except Exception as e:
    print("Error:", e)
